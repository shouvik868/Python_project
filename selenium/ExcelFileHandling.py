import openpyxl
#Reading Excel File
file_path="C:/Users/HP/Desktop/new_framework_2025/Python/BookPractice.xlsx"
#fie->workbook->sheet->rows->cells
workbook=openpyxl.load_workbook(file_path)
sheet=workbook["Sheet1"]
row_count=sheet.max_row
cell_count=sheet.max_column
for r in range(2,row_count+1):
    for c in range(1,cell_count+1):
        print(sheet.cell(r,c).value,end="           ")
    print("")

#Writing Excel File
file_path="C:/Users/HP/Desktop/new_framework_2025/Python/BookPractice.xlsx"
workbook=openpyxl.load_workbook(file_path)
sheet=workbook["Sheet2"]
for r in range(1,4):
    for c in range(1,4):
        sheet.cell(r,c).value=str(r+c)
workbook.save(file_path)#mandetory


